import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import re
from openpyxl import load_workbook
from dateutil import parser as dt_parser
from typing import Iterable, List, Optional, Tuple

# ── Original v3.3 shift start times (for reference when parsing comments) ──
# This dict will be used to look up default start times.
# For custom shifts, we will use the full range.
SHIFT_START_DEFAULTS = {
    "FS": "08:00", "First": "08:00",
    "GS": "09:30", "General": "09:30",
    "SS": "11:30", "Second": "11:30",  # Adjusted from v3.3 based on typical SS
    "NS": "20:30", "Night": "20:30",
}

# ── Full shift ranges for expected duration and correct overnight handling ──
# This will be the primary source for shift parsing, especially if custom shifts are ranges.
# If custom_shift_times only provide a start, we'll combine with these default ends.
DEFAULT_SHIFT_RANGES = {
    "FS": "08:00 - 17:00", "First": "08:00 - 17:00",
    "GS": "09:30 - 18:30", "General": "09:30 - 18:30",
    "SS": "13:00 - 21:30", "Second": "13:00 - 21:30",  # Adjusted from v3.3 SS for a full shift range
    "NS": "20:00 - 08:00", "Night": "20:00 - 08:00",  # Next day
}


def _parse_hhmm(txt: str) -> Optional[datetime]:
    """Parses a string 'HH:MM' into a datetime object (date is arbitrary)."""
    try:
        return datetime.strptime(txt.strip(), "%H:%M")
    except ValueError:
        return None


def _parse_hhmm_range(time_range_str: str) -> Tuple[Optional[datetime], Optional[datetime]]:
    """
    Parses a string 'HH:MM - HH:MM' into a tuple of (start_datetime, end_datetime).
    Handles overnight shifts (end_time < start_time implies next day).
    """
    time_range_str = time_range_str.strip()
    parts = re.split(r'\s*-\s*', time_range_str)

    if len(parts) == 2:
        start_time = _parse_hhmm(parts[0])
        end_time = _parse_hhmm(parts[1])
        if start_time and end_time:
            # If end time is earlier than start time, assume it's on the next day
            if end_time < start_time:
                end_time += timedelta(days=1)
            return start_time, end_time
    return None, None  # Invalid format or single time string (not a range)


# -- Analyze comments/notes logic -----------------------------------

def override_by_comment(note: str, t_in: Optional[datetime], t_out: Optional[datetime]) -> dict:
    """Detect keywords in cell‑comments and override daily status / OT logic."""
    o = {"force_status": None, "skip_half_day": False, "late_override": None, "extra_wh": 0.0}
    if not note:
        return o

    nl = note.strip().lower()

    if nl == "short leave":
        o["force_status"] = "P"
        o["skip_half_day"] = True
        return o

    if nl == "half day":
        o["force_status"] = "0.5P"
        o["skip_half_day"] = True
        return o

    # Combined logic for "out:" comments: prioritize specific HH:MM parsing
    # and also handle "out: and back in:" for extra_wh calculation.
    if "out:" in nl:
        if "back in:" in nl:
            # v3.3 original logic for "out: and back in:"
            times = re.findall(r'(\d{1,2}[:\.]\d{1,2}\s*(?:am|pm)?)', note, re.I)
            parsed_times = []
            for t_str in times:
                try:
                    parsed_times.append(dt_parser.parse(t_str, fuzzy=True))
                except Exception:
                    pass
            if len(parsed_times) >= 2:
                diff = (parsed_times[1] - parsed_times[0]).total_seconds() / 3600
                o["extra_wh"] = diff
            o["force_status"] = "P"
            o["skip_half_day"] = True  # Original v3.3 behavior
        else:
            # More robust 'out:HH:MM' or 'out:HHMM' parsing
            try:
                parts = nl.split("out:")
                out_str = parts[1].strip()
                forced_out = None
                if re.match(r"^\d{1,2}:\d{2}$", out_str):
                    forced_out = datetime.strptime(out_str, "%H:%M")
                elif re.match(r"^\d{3,4}$", out_str):
                    if len(out_str) == 3: out_str = '0' + out_str
                    forced_out = datetime.strptime(out_str, "%H%M")

                if forced_out and t_in:
                    if forced_out < t_in:  # Handle overnight forced out
                        forced_out += timedelta(days=1)
                    o["extra_wh"] = (forced_out - t_in).total_seconds() / 3600
                    o["force_status"] = "P"  # Force presence if out time is recorded
            except Exception:
                pass  # Malformed "out:" comment, ignore

    if nl.startswith("in:"):  # v3.3 original logic
        o["force_status"] = "P"
        o["skip_half_day"] = True
        o["late_override"] = False  # This means late is explicitly overridden to false
        return o  # Return immediately if "in:" is the primary directive

    if "late:" in nl:  # Improved late override logic
        try:
            parts = nl.split("late:")
            late_min_str = parts[1].strip()
            late_minutes = int(late_min_str)
            o["late_override"] = late_minutes
        except Exception:
            pass  # Malformed "late:" comment, ignore

    return o


def build_master(
        shifted_path: Path,  # Changed to Path, assuming gui_app passes a Path object
        save_path: Path,  # Changed to Path
        analyze_comments: bool = False,
        shifts_path: Optional[Path] = None,  # Changed to Path
        ot_filter: Optional[Iterable[str]] = None,  # Keep as Iterable[str]
        custom_shift_times: Optional[dict] = None,  # Custom shift times from GUI (HH:MM - HH:MM)
):
    """
    Builds the final master attendance sheet and an OT sheet.
    Incorporates custom shift times and comment analysis.

    Args:
        shifted_path (Path): Path to the DataFrame after shifts have been added.
        save_path (Path): Path to save the output Excel file.
        analyze_comments (bool): Whether to analyze cell comments in the shifts file.
        shifts_path (Optional[Path]): Path to the original shifts file, required if analyze_comments is True.
        ot_filter (Optional[Iterable[str]]): List of employee IDs to filter for the OT table.
        custom_shift_times (Optional[dict]): Dictionary of custom shift time strings
                                              (e.g., {"FS": "08:00 - 17:00"}).
    """
    # Defensive checks to ensure paths are Path objects if they came as strings
    if not isinstance(shifted_path, Path):
        shifted_path = Path(shifted_path)
    if analyze_comments and shifts_path and not isinstance(shifts_path, Path):
        shifts_path = Path(shifts_path)
    if not isinstance(save_path, Path):
        save_path = Path(save_path)

    if analyze_comments and not shifts_path:
        raise ValueError("shifts_path must be provided if analyze_comments is True.")

    df = pd.read_excel(shifted_path)

    wb = None
    if analyze_comments:
        try:
            wb = load_workbook(shifts_path, data_only=True)
            ws = wb.active
        except Exception as e:
            print(f"Warning: Could not load shifts file for comment analysis: {e}")
            analyze_comments = False

    # ── Prepare shift schedules (combining defaults and custom) ────────────────
    # This will hold the parsed datetime objects for start and end times for each shift code
    parsed_shift_schedules = {}

    # 1. Start with default shift ranges (for full duration calculation)
    for code, time_range_str in DEFAULT_SHIFT_RANGES.items():
        start_dt, end_dt = _parse_hhmm_range(time_range_str)
        if start_dt and end_dt:
            parsed_shift_schedules[code] = {"start": start_dt, "end": end_dt}

    # 2. Override with custom shift times if provided (these are full HH:MM - HH:MM ranges)
    if custom_shift_times:
        for k, v_range_str in custom_shift_times.items():
            start_dt, end_dt = _parse_hhmm_range(v_range_str)
            if start_dt and end_dt:
                parsed_shift_schedules[k] = {"start": start_dt, "end": end_dt}
            elif start_dt:  # If only start is valid, (e.g., custom was just HH:MM)
                # Try to use default end time if a custom start was given but no end
                default_end_dt = parsed_shift_schedules.get(k, {}).get("end")
                if default_end_dt:
                    if default_end_dt < start_dt:
                        default_end_dt += timedelta(days=1)  # Ensure overnight if needed
                    parsed_shift_schedules[k] = {"start": start_dt, "end": default_end_dt}
                else:
                    parsed_shift_schedules[k] = {"start": start_dt, "end": None}  # No end known

    # Fallback default for schedule if a shift code isn't found
    default_sched_start_fallback = _parse_hhmm(SHIFT_START_DEFAULTS["GS"])  # Use General Shift start as default
    default_sched_end_fallback = parsed_shift_schedules.get("GS", {}).get("end") or _parse_hhmm(
        DEFAULT_SHIFT_RANGES["GS"].split('-')[1].strip())
    if default_sched_start_fallback and default_sched_end_fallback and default_sched_end_fallback < default_sched_start_fallback:
        default_sched_end_fallback += timedelta(days=1)

    # Fallback for just shift_start lookup (for late mark calculation if specific shift isn't found)
    default_late_check_start_fallback = _parse_hhmm(SHIFT_START_DEFAULTS["GS"])

    # ---- Extract date columns and first date (v3.3 original logic) ------------------------
    date_cols = df.columns[3:]
    num_days = len(date_cols)

    # Use a consistent year for parsing date headers (e.g., "01-Jan"). Original v3.3 used 2025.
    first_date = datetime.strptime(f"{date_cols[0].strip()}-2025", "%d-%b-%Y")

    # ---- pull comments (optional) (v3.3 original logic) ---------------------------------
    comments_map = {}
    if analyze_comments and wb:
        ws = wb.active
        for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
            emp_id_cell = row[1]
            if emp_id_cell.value is None:
                continue

            emp_id_val = str(int(emp_id_cell.value)) if isinstance(emp_id_cell.value, (int, float)) else str(
                emp_id_cell.value).strip()
            comments_map.setdefault(emp_id_val, {})
            # Loop through cells corresponding to date columns (starting from index 3)
            for col_idx_excel, cell in enumerate(row[3:3 + num_days]):
                if cell.comment and cell.comment.text.strip():
                    comments_map[emp_id_val][col_idx_excel] = cell.comment.text.strip()

    # ---- Iterate 9‑row employee blocks (v3.3 original logic) ----------------------------
    master_rows, ot_rows = [], []
    block_size = 9

    filter_set: Optional[set[str]] = (
        {str(c).strip() for c in ot_filter} if ot_filter else None
    )

    for sr_idx, top in enumerate(range(0, len(df), block_size), start=1):
        block = df.iloc[top: top + block_size]
        if block.empty:
            continue

        emp_id_raw, emp_name_raw = block.iloc[0][["EmpCode", "EmpName"]]
        emp_id = str(int(emp_id_raw)) if isinstance(emp_id_raw, (int, float)) else str(emp_id_raw).strip()
        emp_name = str(emp_name_raw).strip()

        # Extract rows by metric, focusing on the date columns (from index 3 onwards)
        status_row = block.loc[block["metric"] == "Status"].iloc[0, 3:]
        in_row = block.loc[block["metric"] == "InTime"].iloc[0, 3:]
        out_row = block.loc[block["metric"] == "OutTime"].iloc[0, 3:]
        shift_row = block.loc[block["metric"] == "Shift"].iloc[0, 3:]
        late_by_row = block.loc[block["metric"] == "Late By"].iloc[0, 3:]

        present = 0
        leave = 0
        c_off = 0.0
        od1 = 0
        od2 = 0
        ot_hours_total = 0  # Accumulate OT hours for this employee, only if in filter
        late = 0
        total_actual_wh = 0.0  # Total actual working hours for average calculation

        daily_status_list = []
        daily_ots_list = []

        emp_comments_map = comments_map.get(emp_id, {}) if analyze_comments else {}

        # ── per‑date loop ───────────────────────────────────────────
        for day_idx, col_date_str in enumerate(date_cols):
            status = str(status_row.get(col_date_str, "")).strip()

            t_in = _parse_hhmm(str(in_row.get(col_date_str)))
            t_out = _parse_hhmm(str(out_row.get(col_date_str)))

            wh_today = 0.0  # Initialize for current day
            comment_override = {"force_status": None, "skip_half_day": False, "late_override": None, "extra_wh": 0.0}

            if analyze_comments:
                comment_text = emp_comments_map.get(day_idx)
                if comment_text:
                    comment_override = override_by_comment(comment_text, t_in, t_out)

            # Re-calculate t_out if extra_wh is provided by comment
            if comment_override["extra_wh"] > 0 and t_in:
                t_out = t_in + timedelta(hours=comment_override["extra_wh"])
                # Ensure t_out is on the correct day for duration calculation
                if t_out < t_in:  # If adding hours somehow pushed it back (unlikely but defensive)
                    t_out += timedelta(days=1)

            # Calculate working hours (wh_today)
            if t_in and t_out:
                wh_today = (t_out - t_in).total_seconds() / 3600
                # Handle overnight shifts where t_out is effectively next day
                if t_out < t_in:
                    wh_today = ((t_out + timedelta(days=1)) - t_in).total_seconds() / 3600
            else:
                wh_today = 0.0  # No valid in/out times

            # --- Determine Final Status ---
            final_status = status  # Start with existing status from Excel

            if comment_override["force_status"]:  # Comment override takes precedence
                final_status = comment_override["force_status"]
            elif final_status == "P" or final_status == "":  # If initially 'P' or empty, apply rules
                if wh_today >= 7.0:
                    final_status = "P"
                elif 3.0 <= wh_today < 7.0:  # Half day condition
                    if not comment_override["skip_half_day"]:
                        final_status = "0.5P"
                    else:
                        final_status = "P"  # Skip half-day if comment implies (e.g. short leave)
                else:
                    final_status = "A"  # Default to Absent if less than 3 hours

            # Update counters based on final_status
            if final_status == "P":
                present += 1
            elif final_status == "0.5P":
                present += 0.5
            elif final_status == "L":
                present += 1
            elif final_status == "OD1":
                present += 1
                od1 += 1
            elif final_status == "OD2":
                present += 1
                od2 += 1
            elif final_status == "Leave":
                leave += 1
            elif final_status == "C-off":
                c_off += 1

            # --- Late Mark Calculation ---
            late_flag = False
            if final_status in ["P", "0.5P", "L", "OD1", "OD2"] and t_in:  # Only check late if present and has in time
                shift_code = str(shift_row.get(col_date_str, "")).strip()

                # Get scheduled start time: prioritize parsed_shift_schedules, then SHIFT_START_DEFAULTS
                scheduled_start_for_late_check = parsed_shift_schedules.get(shift_code, {}).get("start")
                if not scheduled_start_for_late_check:  # Fallback to SHIFT_START_DEFAULTS for single HH:MM
                    sched_str = SHIFT_START_DEFAULTS.get(shift_code, SHIFT_START_DEFAULTS["GS"])
                    scheduled_start_for_late_check = _parse_hhmm(sched_str)

                if scheduled_start_for_late_check:
                    late_threshold_minutes = 15  # Default
                    if comment_override["late_override"] is not None:
                        late_threshold_minutes = comment_override["late_override"]

                    if (t_in - scheduled_start_for_late_check).total_seconds() / 60 > late_threshold_minutes:
                        late_flag = True
                else:  # If no scheduled start time found, try using Late By column as fallback for 'late' status
                    late_txt = str(late_by_row.get(col_date_str, "")).strip()
                    if late_txt and late_txt != "00:00":
                        lt_dt = _parse_hhmm(late_txt)
                        if lt_dt and (lt_dt.hour * 60 + lt_dt.minute) > 15:
                            late_flag = True

            if late_flag:
                late += 1
                if final_status == "P":  # Change 'P' to 'L' if late
                    final_status = "L"

            daily_status_list.append(final_status)

            # --- Overtime (OT) and C-off logic (v3.3 logic) -------------------------
            ot_today = 0
            if wh_today > 0:  # Only count if there was actual working time
                total_actual_wh += wh_today  # For average calculation

                extra_hours = wh_today - 8.5  # OT calculation based on 8.5 hours working day

                if (filter_set is None) or (emp_id in filter_set):
                    # Calculate OT only for filtered people
                    if extra_hours > 0:
                        raw_ot = extra_hours
                        ot_today = int(raw_ot) + (1 if (raw_ot - int(raw_ot)) >= 0.75 else 0)
                        ot_hours_total += ot_today
                else:
                    # Calculate c-off only for non-OT people
                    if 3.5 <= extra_hours < 4.0:
                        c_off += 0.5
                    elif extra_hours >= 7.0:
                        c_off += 1.0

            daily_ots_list.append(f"{ot_today:.2f}")  # Format OT to 2 decimal places

        # ── End of per‑date loop ────────────────────────────────────

        avg_wh = round(total_actual_wh / num_days, 2) if num_days > 0 else 0  # Average for all days in the period

        master_rows.append([
            sr_idx + 1, emp_id, emp_name, *daily_status_list,
            present, leave, 5, 0, num_days,  # 5 W. off, 0 Holiday (hardcoded as per original)
            c_off, 20, 19, od1, od2, avg_wh, ot_hours_total, late, "", ""  # 20 TA adjusted, 19 TA (hardcoded)
        ])

        # -- OT table row (filter aware) ------------------------------
        if (filter_set is None) or (emp_id in filter_set):
            ot_rows.append([sr_idx + 1, emp_id, emp_name, *daily_ots_list, ot_hours_total])

    # ---- build DataFrames ------------------------------------------
    day_headers = [d.strftime("%d") for d in pd.date_range(first_date, periods=num_days)]

    master_cols = (
            ["Sr. no.", "Emp. ID", "Emp. name"] + day_headers +
            ["Present", "Leave", "W. off", "Holiday", "Total", "C-off",
             "TA adjusted", "TA", "OD1", "OD2", "Avg Working Hr",
             "OT", "Late marks", "Leave cut", "Remarks"]
    )

    df_master = pd.DataFrame(master_rows, columns=master_cols)
    df_master.index += 1  # start index at 1 visually (optional)

    ot_cols = ["Sr. no.", "Emp. ID", "Emp. name"] + day_headers + ["Total OT"]
    df_ot = pd.DataFrame(ot_rows, columns=ot_cols) if ot_rows else pd.DataFrame(columns=ot_cols)

    # ---- write both tables with 5‑row spacing (v3.3 original logic) -----------------------
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        df_master.to_excel(writer, index=False, sheet_name="Sheet1", startrow=0)
        if not df_ot.empty:
            startrow_ot = len(df_master) + 6  # +1 for header, +5 for spacing
            df_ot.to_excel(writer, index=False, sheet_name="Sheet1", startrow=startrow_ot)

    print(f"✅ MasterSheet + OT table saved → {save_path}")